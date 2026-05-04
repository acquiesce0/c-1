import os
import re
import subprocess
import sys
import threading
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import openpyxl

DATA_ROOT = Path(__file__).parent.resolve()
YEARS_DIR = DATA_ROOT / "years"
TOLERANCE = 0.0005
FILENAME_RE = re.compile(r"(?i)pans?\s*(\d{1,2})-(\d{4})")


def parse_filename(name):
    m = FILENAME_RE.search(name)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def label_kind(cell):
    if cell is None:
        return None
    s = str(cell).strip().lower()
    if not s:
        return None
    if "density" in s:
        return "density"
    if s.startswith("mg"):
        return "mg"
    if s.startswith("ca"):
        return "ca"
    if s.startswith("na"):
        return "na"
    if s.startswith("k"):
        if "stpb" in s:
            return None
        if "a.a" in s or "aa" in s:
            return "k"
    return None


def to_float(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def normalize_name(s):
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).upper()


def name_match(query, target):
    q = normalize_name(query)
    t = normalize_name(target)
    if not q or not t:
        return False
    if q == t:
        return True
    tokens = [seg.strip() for seg in t.split("/") if seg.strip()]
    return q in tokens


def extract_records(filepath):
    parsed = parse_filename(filepath.name)
    if not parsed:
        return []
    month_hint, year_hint = parsed
    records = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception:
        return []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            day_digits = re.sub(r"\D", "", sheet_name)
            if not day_digits:
                continue
            day_hint = int(day_digits)
            if not 1 <= day_hint <= 31:
                continue
            try:
                row_data = {}
                for i, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=30, values_only=True), 1
                ):
                    row_data[i] = row
            except Exception:
                continue

            labels = {}
            for r, row in row_data.items():
                if not row:
                    continue
                kind = label_kind(row[0])
                if kind and kind not in labels:
                    labels[kind] = r
            if "density" not in labels:
                continue
            density_row_idx = labels["density"]

            header_row_idx = None
            for r in range(max(1, density_row_idx - 6), density_row_idx):
                row = row_data.get(r) or ()
                text_count = sum(
                    1 for v in row[1:] if isinstance(v, str) and v.strip()
                )
                if text_count >= 4:
                    header_row_idx = r
            if header_row_idx is None:
                continue

            headers = row_data[header_row_idx]
            density_vals = row_data[density_row_idx]
            mg_vals = row_data.get(labels.get("mg", 0)) or ()
            ca_vals = row_data.get(labels.get("ca", 0)) or ()
            k_vals = row_data.get(labels.get("k", 0)) or ()
            na_vals = row_data.get(labels.get("na", 0)) or ()

            sample_date = None
            for r, row in row_data.items():
                if not row:
                    continue
                if isinstance(row[0], str) and "sampling" in row[0].lower():
                    for v in row[1:6]:
                        if isinstance(v, datetime):
                            sample_date = v.date()
                            break
                    break
            if sample_date is None:
                try:
                    sample_date = datetime(year_hint, month_hint, day_hint).date()
                except ValueError:
                    continue

            def cell(arr, idx):
                if idx < len(arr):
                    return arr[idx]
                return None

            for ci in range(1, len(headers)):
                name = headers[ci]
                if not isinstance(name, str) or not name.strip():
                    continue
                density = to_float(cell(density_vals, ci))
                if density is None:
                    continue
                records.append(
                    {
                        "file": filepath.name,
                        "filepath": str(filepath),
                        "sheet": sheet_name,
                        "year": year_hint,
                        "month": month_hint,
                        "day": day_hint,
                        "date": sample_date,
                        "location": name.strip(),
                        "density": density,
                        "mg": to_float(cell(mg_vals, ci)),
                        "ca": to_float(cell(ca_vals, ci)),
                        "k": to_float(cell(k_vals, ci)),
                        "na": to_float(cell(na_vals, ci)),
                    }
                )

            lower_hdr = None
            for r, row in row_data.items():
                if not row:
                    continue
                for c, v in enumerate(row):
                    if isinstance(v, str) and "density @ 45" in v.lower():
                        lower_hdr = (r, c)
                        break
                if lower_hdr:
                    break
            if lower_hdr:
                hr, hc = lower_hdr
                hdr_row = row_data.get(hr) or ()
                mg_at_plus2 = (
                    hc + 2 < len(hdr_row)
                    and isinstance(hdr_row[hc + 2], str)
                    and hdr_row[hc + 2].strip().lower().startswith("mg")
                )
                if mg_at_plus2:
                    for dr in range(hr + 1, hr + 9):
                        row = row_data.get(dr) or ()
                        if hc >= len(row):
                            continue
                        label = row[hc]
                        if label is None:
                            continue
                        if isinstance(label, str):
                            name = label.strip()
                            if not name:
                                continue
                        else:
                            name = str(label).strip()
                            if not name:
                                continue
                        density = to_float(cell(row, hc + 1))
                        if density is None:
                            continue
                        records.append(
                            {
                                "file": filepath.name,
                                "filepath": str(filepath),
                                "sheet": sheet_name,
                                "year": year_hint,
                                "month": month_hint,
                                "day": day_hint,
                                "date": sample_date,
                                "location": name,
                                "density": density,
                                "mg": to_float(cell(row, hc + 2)),
                                "ca": to_float(cell(row, hc + 3)),
                                "k": to_float(cell(row, hc + 4)),
                                "na": to_float(cell(row, hc + 5)),
                            }
                        )
    finally:
        wb.close()
    return records


def find_data_files():
    if not YEARS_DIR.is_dir():
        return []
    files = []
    for sub in YEARS_DIR.iterdir():
        if not sub.is_dir():
            continue
        for f in sub.iterdir():
            if f.suffix.lower() == ".xlsx" and FILENAME_RE.search(f.name):
                files.append(f)
    return sorted(files)


def load_all_records(progress_cb=None):
    files = find_data_files()
    all_recs = []
    for i, f in enumerate(files):
        if progress_cb:
            progress_cb(i + 1, len(files), f.name)
        all_recs.extend(extract_records(f))
    return all_recs


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Brine Density Lookup")
        self.geometry("1180x650")
        self.records = []
        self.last_results = []
        self.last_query = None
        self._row_records = {}
        self._date_row_records = {}
        self.last_date_results = []
        self.last_date_query = None
        self._build_ui()
        self.after(50, self.load_data)

    def _build_ui(self):
        self.status_var = tk.StringVar(value="Loading data...")

        status_bar = ttk.Frame(self, relief="groove", borderwidth=1)
        status_bar.pack(side="bottom", fill="x")
        self.progress = ttk.Progressbar(
            status_bar, orient="horizontal", mode="determinate", length=260,
        )
        self.progress.pack(side="right", padx=8, pady=4)
        ttk.Label(
            status_bar, textvariable=self.status_var, anchor="w",
            font=("Segoe UI", 9),
        ).pack(side="left", fill="x", expand=True, padx=10, pady=4)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(side="top", fill="both", expand=True, padx=10, pady=(10, 4))

        self.search_tab = ttk.Frame(self.notebook)
        self.date_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.search_tab, text="Material Search")
        self.notebook.add(self.date_tab, text="Date Comparison")

        self._build_search_tab(self.search_tab)
        self._build_date_tab(self.date_tab)
        self.bind("<Return>", self._on_enter_pressed)
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

    def _on_tab_changed(self, _event=None):
        try:
            idx = self.notebook.index("current")
        except tk.TclError:
            return
        if idx == 1:
            target = (1420, 780)
        else:
            target = (1180, 650)
        cur_w = self.winfo_width()
        cur_h = self.winfo_height()
        new_w = max(cur_w, target[0])
        new_h = max(cur_h, target[1])
        if new_w != cur_w or new_h != cur_h:
            self.geometry(f"{new_w}x{new_h}")

    def _on_enter_pressed(self, _event):
        if self.notebook.index("current") == 0:
            self.search()
        else:
            self.date_search()

    def _build_search_tab(self, parent):
        top = ttk.Frame(parent, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="Material name:").grid(row=0, column=0, sticky="w", padx=4)
        self.name_var = tk.StringVar()
        self.name_var.trace_add("write", self._uppercase_name)
        self.name_entry = ttk.Entry(
            top, textvariable=self.name_var, width=18,
            font=("Segoe UI", 11, "bold"),
        )
        self.name_entry.grid(row=0, column=1, padx=4)

        ttk.Label(top, text="Density:").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.density_var = tk.StringVar()
        self.density_entry = ttk.Entry(
            top, textvariable=self.density_var, width=12,
            font=("Segoe UI", 11, "bold"),
        )
        self.density_entry.grid(row=0, column=3, padx=4)
        ttk.Label(top, text=f"(±{TOLERANCE})").grid(row=0, column=4, sticky="w")

        self.search_btn = ttk.Button(top, text="Search", command=self.search)
        self.search_btn.grid(row=0, column=5, padx=8)
        self.reload_btn = ttk.Button(top, text="Reload data", command=self.load_data)
        self.reload_btn.grid(row=0, column=6, padx=4)
        self.export_btn = ttk.Button(
            top, text="Export to Excel", command=self.export_results, state="disabled"
        )
        self.export_btn.grid(row=0, column=7, padx=4)

        body = ttk.Frame(parent)
        body.pack(fill="both", expand=True, padx=10, pady=10)

        cols = (
            "year", "month", "day", "location", "density",
            "mg", "ca", "k", "na", "file",
        )
        headings = {
            "year": "Year", "month": "Month", "day": "Day",
            "location": "Material", "density": "Density",
            "mg": "Mg+2", "ca": "Ca+2", "k": "K+ A.A.", "na": "Na+ A.A.",
            "file": "Source File",
        }
        widths = {
            "year": 60, "month": 60, "day": 50, "location": 110, "density": 80,
            "mg": 90, "ca": 90, "k": 90, "na": 90, "file": 220,
        }
        self.tree = ttk.Treeview(body, columns=cols, show="headings", height=22)
        for c in cols:
            self.tree.heading(c, text=headings[c])
            self.tree.column(
                c, width=widths[c], anchor="center" if c != "file" else "w"
            )
        self.tree.tag_configure("group", background="#e8eef7", font=("Segoe UI", 9, "bold"))
        self.tree.bind("<Double-1>", self._on_row_double_click)
        sb = ttk.Scrollbar(body, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="left", fill="y")

    def _build_date_tab(self, parent):
        top = ttk.Frame(parent, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="Day:").grid(row=0, column=0, sticky="w", padx=4)
        self.day_var = tk.StringVar()
        ttk.Entry(
            top, textvariable=self.day_var, width=5,
            font=("Segoe UI", 11, "bold"),
        ).grid(row=0, column=1, padx=4)

        ttk.Label(top, text="Month:").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.month_var = tk.StringVar()
        ttk.Entry(
            top, textvariable=self.month_var, width=5,
            font=("Segoe UI", 11, "bold"),
        ).grid(row=0, column=3, padx=4)

        ttk.Label(top, text="Year:").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.year_var = tk.StringVar()
        ttk.Entry(
            top, textvariable=self.year_var, width=8,
            font=("Segoe UI", 11, "bold"),
        ).grid(row=0, column=5, padx=4)

        self.date_search_btn = ttk.Button(top, text="Compare", command=self.date_search)
        self.date_search_btn.grid(row=0, column=6, padx=8)
        self.date_export_btn = ttk.Button(
            top, text="Export to Excel",
            command=self.export_date_results, state="disabled",
        )
        self.date_export_btn.grid(row=0, column=7, padx=4)

        ttk.Label(
            top,
            text="Pick a specific spreadsheet (day/month/year). "
                 "Top pane shows that day; bottom shows every other record "
                 "in the folder for the same materials, with Δ vs. the selected day.",
            foreground="#555",
            wraplength=900,
            justify="left",
        ).grid(row=1, column=0, columnspan=8, sticky="w", padx=4, pady=(4, 0))

        body = ttk.PanedWindow(parent, orient="vertical")
        body.pack(fill="both", expand=True, padx=10, pady=10)

        sel_frame = ttk.LabelFrame(body, text="Selected day")
        body.add(sel_frame, weight=1)

        sel_cols = (
            "material", "density", "mg", "ca", "k", "na", "file",
        )
        sel_headings = {
            "material": "Material", "density": "Density",
            "mg": "Mg+2", "ca": "Ca+2", "k": "K+ A.A.", "na": "Na+ A.A.",
            "file": "Source File",
        }
        sel_widths = {
            "material": 140, "density": 90,
            "mg": 90, "ca": 90, "k": 90, "na": 90, "file": 240,
        }
        self.sel_tree = ttk.Treeview(
            sel_frame, columns=sel_cols, show="headings", height=8
        )
        for c in sel_cols:
            self.sel_tree.heading(c, text=sel_headings[c])
            self.sel_tree.column(
                c, width=sel_widths[c],
                anchor="center" if c != "file" else "w",
            )
        self.sel_tree.bind("<Double-1>", self._on_sel_row_double_click)
        sel_sb = ttk.Scrollbar(sel_frame, orient="vertical", command=self.sel_tree.yview)
        self.sel_tree.configure(yscrollcommand=sel_sb.set)
        self.sel_tree.pack(side="left", fill="both", expand=True)
        sel_sb.pack(side="left", fill="y")

        cmp_frame = ttk.LabelFrame(
            body,
            text=f"Comparison — same material, density ±{TOLERANCE}",
        )
        body.add(cmp_frame, weight=2)

        cols = (
            "material", "year", "date", "density",
            "mg", "ca", "k", "na", "file",
        )
        headings = {
            "material": "Material", "year": "Year", "date": "Date",
            "density": "Density", "mg": "Mg+2", "ca": "Ca+2",
            "k": "K+ A.A.", "na": "Na+ A.A.",
            "file": "Source File",
        }
        widths = {
            "material": 130, "year": 60, "date": 100,
            "density": 90, "mg": 90, "ca": 90, "k": 90, "na": 90,
            "file": 240,
        }
        self.date_tree = ttk.Treeview(cmp_frame, columns=cols, show="headings", height=14)
        for c in cols:
            self.date_tree.heading(c, text=headings[c])
            self.date_tree.column(
                c, width=widths[c], anchor="center" if c != "file" else "w"
            )
        self.date_tree.tag_configure(
            "group", background="#e8eef7", font=("Segoe UI", 9, "bold")
        )
        self.date_tree.bind("<Double-1>", self._on_date_row_double_click)
        sb = ttk.Scrollbar(cmp_frame, orient="vertical", command=self.date_tree.yview)
        self.date_tree.configure(yscrollcommand=sb.set)
        self.date_tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="left", fill="y")

        self._sel_row_records = {}

    def _uppercase_name(self, *_):
        current = self.name_var.get()
        upper = current.upper()
        if current != upper:
            self.name_var.set(upper)

    def load_data(self):
        self.search_btn.config(state="disabled")
        self.reload_btn.config(state="disabled")
        self.date_search_btn.config(state="disabled")

        self.status_var.set("Scanning years/ folder...")
        self.progress.config(mode="indeterminate", value=0)
        self.progress.start(80)
        self.update_idletasks()

        def progress(i, total, name):
            def update():
                if total > 0:
                    if str(self.progress["mode"]) != "determinate":
                        self.progress.stop()
                        self.progress.config(mode="determinate", maximum=total)
                    self.progress["value"] = i
                self.status_var.set(f"Loading {i}/{total}: {name}")
            self.after(0, update)

        def worker():
            try:
                recs = load_all_records(progress)
            except Exception as e:
                self.after(
                    0,
                    lambda e=e: messagebox.showerror("Load failed", str(e)),
                )
                recs = []
            self.after(0, self._on_loaded, recs)

        threading.Thread(target=worker, daemon=True).start()

    def _on_loaded(self, recs):
        try:
            self.progress.stop()
        except Exception:
            pass
        self.progress.config(mode="determinate", value=0, maximum=100)
        self.records = recs
        self.search_btn.config(state="normal")
        self.reload_btn.config(state="normal")
        self.date_search_btn.config(state="normal")
        self.name_entry.focus_set()
        if not YEARS_DIR.is_dir():
            self.status_var.set(
                f"No 'years' folder found at {YEARS_DIR}. "
                "Create it and place year subfolders (e.g. 2027) inside, "
                "each containing 'Pan MM-YYYY.xlsx' files."
            )
            return
        if not recs:
            self.status_var.set(
                f"'years' folder is empty or contains no recognizable files. "
                f"Add a year subfolder under {YEARS_DIR} with 'Pan MM-YYYY.xlsx' files, "
                "then click Reload data."
            )
            return
        years = sorted({r["year"] for r in recs})
        years_str = ", ".join(str(y) for y in years)
        self.status_var.set(
            f"Loaded {len(recs)} records from years: {years_str}. "
            f"Enter a material name and density, then press Search (or Enter)."
        )

    def search(self):
        name = self.name_var.get().strip()
        d_str = self.density_var.get().strip()
        if not name:
            messagebox.showwarning("Missing input", "Enter a material name.")
            return
        try:
            density = float(d_str)
        except ValueError:
            messagebox.showwarning(
                "Invalid density", f"'{d_str}' is not a number."
            )
            return

        for iid in self.tree.get_children():
            self.tree.delete(iid)
        self._row_records.clear()

        self.status_var.set(f"Searching for '{name}' at density {density:.4f}...")
        self.progress.config(mode="indeterminate")
        self.progress.start(60)
        self.update_idletasks()

        results = []
        for r in self.records:
            if not name_match(name, r["location"]):
                continue
            if abs(r["density"] - density) > TOLERANCE:
                continue
            results.append(r)

        self.progress.stop()
        self.progress.config(mode="determinate", value=0, maximum=100)

        self.last_results = results
        self.last_query = (name, density)
        self.export_btn.config(state="normal" if results else "disabled")

        if not results:
            self.status_var.set(
                f"No matches for '{name}' near density {density:.4f} (±{TOLERANCE})."
            )
            return

        results.sort(
            key=lambda r: (r["year"], r["month"], r["day"], r["location"])
        )
        per_year = {}
        for r in results:
            per_year.setdefault(r["year"], []).append(r)

        def fmt(v):
            return f"{v:.4f}" if isinstance(v, (int, float)) else ""

        for year in sorted(per_year):
            recs = per_year[year]
            self.tree.insert(
                "",
                "end",
                values=(
                    f"— {year} —",
                    f"{len(recs)} match(es)",
                    "", "", "", "", "", "", "", "",
                ),
                tags=("group",),
            )
            for r in recs:
                iid = self.tree.insert(
                    "",
                    "end",
                    values=(
                        r["year"], r["month"], r["day"], r["location"],
                        fmt(r["density"]), fmt(r["mg"]), fmt(r["ca"]),
                        fmt(r["k"]), fmt(r["na"]), r["file"],
                    ),
                )
                self._row_records[iid] = r

        years_summary = ", ".join(
            f"{y}: {len(recs)}" for y, recs in sorted(per_year.items())
        )
        self.status_var.set(
            f"Found {len(results)} match(es) for '{name}' at "
            f"{density:.4f} ±{TOLERANCE} — {years_summary}"
        )

    def export_results(self):
        if not self.last_results:
            messagebox.showinfo("Nothing to export", "Run a search first.")
            return

        name, density = self.last_query
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", name).strip("_") or "results"
        default_filename = f"brine_{safe_name}_{density:.4f}.xlsx"

        path = filedialog.asksaveasfilename(
            title="Export results",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=default_filename,
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        headers = [
            "Year", "Month", "Day", "Date", "Material", "Density",
            "Mg+2", "Ca+2", "K+ A.A.", "Na+ A.A.", "Source File",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        results = sorted(
            self.last_results,
            key=lambda r: (r["year"], r["month"], r["day"], r["location"]),
        )
        for r in results:
            ws.append([
                r["year"], r["month"], r["day"],
                r["date"].isoformat() if r.get("date") else "",
                r["location"], r["density"],
                r["mg"], r["ca"], r["k"], r["na"], r["file"],
            ])

        widths = [6, 7, 5, 12, 16, 10, 10, 10, 10, 10, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        try:
            wb.save(path)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return
        finally:
            wb.close()

        self.status_var.set(f"Exported {len(results)} record(s) to {path}")

    def date_search(self):
        d_str = self.day_var.get().strip()
        m_str = self.month_var.get().strip()
        y_str = self.year_var.get().strip()
        if not d_str or not m_str or not y_str:
            messagebox.showwarning(
                "Missing input", "Enter Day, Month, and Year."
            )
            return
        try:
            day = int(d_str)
            month = int(m_str)
            year = int(y_str)
        except ValueError:
            messagebox.showwarning(
                "Invalid input", "Day, Month, and Year must be numbers."
            )
            return
        if not (1 <= day <= 31) or not (1 <= month <= 12):
            messagebox.showwarning(
                "Invalid date", "Day must be 1-31 and Month must be 1-12."
            )
            return

        for iid in self.sel_tree.get_children():
            self.sel_tree.delete(iid)
        for iid in self.date_tree.get_children():
            self.date_tree.delete(iid)
        self._sel_row_records.clear()
        self._date_row_records.clear()

        self.status_var.set(
            f"Comparing {day:02d}/{month:02d}/{year} against the folder..."
        )
        self.progress.config(mode="indeterminate")
        self.progress.start(60)
        self.update_idletasks()

        selected = [
            r for r in self.records
            if r["day"] == day and r["month"] == month and r["year"] == year
        ]

        if not selected:
            self.progress.stop()
            self.progress.config(mode="determinate", value=0, maximum=100)
            self.last_date_results = []
            self.last_date_query = None
            self.date_export_btn.config(state="disabled")
            self.status_var.set(
                f"No spreadsheet found for {day:02d}/{month:02d}/{year}."
            )
            return

        def fmt(v):
            return f"{v:.4f}" if isinstance(v, (int, float)) else ""

        selected.sort(key=lambda r: r["location"].upper())
        for r in selected:
            iid = self.sel_tree.insert(
                "", "end",
                values=(
                    r["location"],
                    fmt(r["density"]), fmt(r["mg"]), fmt(r["ca"]),
                    fmt(r["k"]), fmt(r["na"]), r["file"],
                ),
            )
            self._sel_row_records[iid] = r

        baseline_by_material = {}
        for r in selected:
            baseline_by_material.setdefault(r["location"].upper(), r)

        compare_records = []
        for r in self.records:
            key = r["location"].upper()
            baseline = baseline_by_material.get(key)
            if baseline is None:
                continue
            if r["day"] == day and r["month"] == month and r["year"] == year:
                continue
            if abs(r["density"] - baseline["density"]) > TOLERANCE:
                continue
            compare_records.append(r)

        by_material = {}
        for r in compare_records:
            by_material.setdefault(r["location"].upper(), []).append(r)

        total_rows = 0
        for material in sorted(baseline_by_material):
            recs = sorted(
                by_material.get(material, []),
                key=lambda r: (r["year"], r["month"], r["day"]),
            )
            self.date_tree.insert(
                "", "end",
                values=(
                    f"— {material} —",
                    f"{len(recs)} match(es)",
                    "", "", "", "", "", "", "",
                ),
                tags=("group",),
            )
            for r in recs:
                date_str = (
                    r["date"].isoformat() if r.get("date") else
                    f"{r['year']}-{r['month']:02d}-{r['day']:02d}"
                )
                iid = self.date_tree.insert(
                    "", "end",
                    values=(
                        r["location"], r["year"], date_str,
                        fmt(r["density"]), fmt(r["mg"]), fmt(r["ca"]),
                        fmt(r["k"]), fmt(r["na"]),
                        r["file"],
                    ),
                )
                self._date_row_records[iid] = r
                total_rows += 1

        self.progress.stop()
        self.progress.config(mode="determinate", value=0, maximum=100)

        self.last_date_results = {
            "selected": selected,
            "compare": compare_records,
            "baselines": baseline_by_material,
        }
        self.last_date_query = (day, month, year)
        self.date_export_btn.config(state="normal")

        self.status_var.set(
            f"Selected {day:02d}/{month:02d}/{year}: "
            f"{len(selected)} material(s). "
            f"Found {total_rows} match(es) in folder with same material "
            f"and density ±{TOLERANCE}."
        )

    def export_date_results(self):
        if not self.last_date_results:
            messagebox.showinfo("Nothing to export", "Run a date comparison first.")
            return

        day, month, year = self.last_date_query
        default_filename = f"compare_{year}-{month:02d}-{day:02d}.xlsx"

        path = filedialog.asksaveasfilename(
            title="Export date comparison",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=default_filename,
        )
        if not path:
            return

        selected = self.last_date_results["selected"]
        compare = self.last_date_results["compare"]
        baselines = self.last_date_results["baselines"]

        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Selected day"
        ws1.append([
            "Material", "Date", "Density", "Mg+2", "Ca+2",
            "K+ A.A.", "Na+ A.A.", "Source File",
        ])
        for cell in ws1[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for r in selected:
            date_str = (
                r["date"].isoformat() if r.get("date") else
                f"{r['year']}-{r['month']:02d}-{r['day']:02d}"
            )
            ws1.append([
                r["location"], date_str,
                r["density"], r["mg"], r["ca"], r["k"], r["na"], r["file"],
            ])
        for i, w in enumerate([18, 12, 10, 10, 10, 10, 10, 30], 1):
            ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws1.freeze_panes = "A2"

        ws2 = wb.create_sheet("Comparison")
        ws2.append([
            "Material", "Year", "Date", "Density", "Mg+2", "Ca+2",
            "K+ A.A.", "Na+ A.A.", "Source File",
        ])
        for cell in ws2[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        by_material = {}
        for r in compare:
            by_material.setdefault(r["location"].upper(), []).append(r)

        for material in sorted(baselines):
            recs = sorted(
                by_material.get(material, []),
                key=lambda r: (r["year"], r["month"], r["day"]),
            )
            for r in recs:
                date_str = (
                    r["date"].isoformat() if r.get("date") else
                    f"{r['year']}-{r['month']:02d}-{r['day']:02d}"
                )
                ws2.append([
                    r["location"], r["year"], date_str,
                    r["density"], r["mg"], r["ca"], r["k"], r["na"],
                    r["file"],
                ])

        for i, w in enumerate(
            [16, 7, 12, 10, 10, 10, 10, 10, 30], 1
        ):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws2.freeze_panes = "A2"

        try:
            wb.save(path)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return
        finally:
            wb.close()

        self.status_var.set(f"Exported date comparison to {path}")

    def _on_sel_row_double_click(self, event):
        iid = self.sel_tree.identify_row(event.y)
        if not iid:
            return
        record = self._sel_row_records.get(iid)
        if record is None:
            return
        filepath = record.get("filepath")
        sheet = record.get("sheet")
        if not filepath:
            return
        opened_at_sheet = self._open_excel_at_sheet(filepath, sheet)
        if opened_at_sheet:
            self.status_var.set(f"Opened {Path(filepath).name} → sheet '{sheet}'.")
        else:
            self.status_var.set(
                f"Opened {Path(filepath).name}. Switch to sheet '{sheet}' manually."
            )

    def _on_date_row_double_click(self, event):
        iid = self.date_tree.identify_row(event.y)
        if not iid:
            return
        record = self._date_row_records.get(iid)
        if record is None:
            return
        filepath = record.get("filepath")
        sheet = record.get("sheet")
        if not filepath:
            return
        opened_at_sheet = self._open_excel_at_sheet(filepath, sheet)
        if opened_at_sheet:
            self.status_var.set(f"Opened {Path(filepath).name} → sheet '{sheet}'.")
        else:
            self.status_var.set(
                f"Opened {Path(filepath).name}. Switch to sheet '{sheet}' manually."
            )

    def _on_row_double_click(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        record = self._row_records.get(iid)
        if record is None:
            return
        filepath = record.get("filepath")
        sheet = record.get("sheet")
        if not filepath:
            return
        opened_at_sheet = self._open_excel_at_sheet(filepath, sheet)
        if opened_at_sheet:
            self.status_var.set(f"Opened {Path(filepath).name} → sheet '{sheet}'.")
        else:
            self.status_var.set(
                f"Opened {Path(filepath).name}. Switch to sheet '{sheet}' manually."
            )

    def _open_excel_at_sheet(self, filepath, sheet_name):
        path = Path(filepath)
        if not path.is_file():
            messagebox.showerror("File not found", f"Could not find:\n{filepath}")
            return False
        try:
            import win32com.client  # type: ignore
        except Exception:
            try:
                if sys.platform == "win32":
                    os.startfile(str(path))
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", str(path)])
                else:
                    subprocess.Popen(["xdg-open", str(path)])
            except Exception as e:
                messagebox.showerror("Open failed", str(e))
            return False
        try:
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
            except Exception:
                excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = True
            target_wb = None
            target_full = str(path.resolve()).lower()
            for wb in excel.Workbooks:
                try:
                    if str(Path(wb.FullName).resolve()).lower() == target_full:
                        target_wb = wb
                        break
                except Exception:
                    continue
            if target_wb is None:
                target_wb = excel.Workbooks.Open(str(path.resolve()))
            target_wb.Activate()
            if sheet_name:
                for sheet in target_wb.Sheets:
                    if sheet.Name == sheet_name:
                        sheet.Activate()
                        try:
                            excel.WindowState = -4137  # xlMaximized
                        except Exception:
                            pass
                        return True
            return False
        except Exception as e:
            messagebox.showerror("Open failed", str(e))
            return False


if __name__ == "__main__":
    App().mainloop()
